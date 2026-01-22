# raporlar/sinav_programi_excel.py
from __future__ import annotations
from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

def programi_xlsx_yaz(dosya_yolu: str, program_kayitlari: List[Dict[str, Any]], sinav_turu: str = "vize"):
    """
    program_kayitlari örnek eleman:
      {
        "bolum": "Bilgisayar Müh.",
        "tarih": date,
        "saat": "09:00",
        "bas": datetime, "bit": datetime,
        "kod": "CSE301",
        "ad": "Algoritmalar",
        "hoca": "Dr. X",
        "derslikler": ["3001","3003"]
      }
    """
    p = Path(dosya_yolu)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sınav Programı"

    # Başlık
    ws.merge_cells("A1:H1")
    ws["A1"] = f"SINAV PROGRAMI ({sinav_turu.upper()})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Bölüm", "Tarih", "Saat", "Ders Kodu", "Ders Adı", "Öğr. Üyesi", "Derslik(ler)", "Bitiş Saati"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # Sıralama: tarih + saat + kod
    rows = sorted(program_kayitlari, key=lambda r: (r.get("tarih"), r.get("saat", ""), r.get("kod", "")))
    for r in rows:
        bit_txt = ""
        try:
            bit_dt = r.get("bit")
            if isinstance(bit_dt, datetime):
                bit_txt = bit_dt.strftime("%H:%M")
        except Exception:
            bit_txt = ""
        ws.append([
            r.get("bolum", ""),
            r.get("tarih", ""),
            r.get("saat", ""),
            r.get("kod", ""),
            r.get("ad", ""),
            r.get("hoca", ""),
            ", ".join(r.get("derslikler", []) or []),
            bit_txt
        ])

    # Sütun genişlikleri
    widths = [20, 12, 10, 12, 34, 24, 18, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Kenarlık
    thin = Side(style="thin")
    last_row = ws.max_row
    for r in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=8):
        for cell in r:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if cell.column in (2, 3, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p.as_posix())
